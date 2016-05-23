#!/usr/bin/python

import sys
import datetime

from openpyxl import load_workbook

from search.models import Variant, Varinfo, pubinfo, indinfo, geneinfo, mosaic
from get_gene_info import get_info 

def strbln(string):
    return string if string else ''

def read_table(xls,sn):
    infos = []
    record = []
    for row in xls[sn].rows:
	for cell in row:
            record.append(strbln(cell.value))
        infos.append(record)
        record=[]
    return infos

def intnone(string):
    return bool(string) and int(string) or None

def floatnone(string):
    return bool(string) and float(string) or None

def intzero(string):
    return bool(string) and int(string) or 0

def invar(f):
    num = 0
    err = ''
    for i in range(len(f)):
        varid, indid, entrez, pmid, gene, chrom, start, end, dna_ref_nt, dna_alt_nt, hgvs, genome_assembly, exon_intron, exon_number, exon_nc, protein_position, pro_ref_aa, pro_alt_aa, frameshift, aa_indel, cdna_position, cdna_ref_aa, cdna_alt_aa, nt_indel, mrna_accession, mrna_length, ref_length, af_lower, af_upper, total_red, sample_type, method, disease = f[i][0:33]
        sys.stderr.write(str(num)+":"+str(entrez)+":"+str(gene)+'\n')
        try:
            var = Varinfo.objects.get(hgvs = hgvs)
        except:
            var = None
        try:
            indidi = indinfo.objects.get(indid=str(indid))
        except:
            err += 'Error: variant id: '+str(varid)+'\n'
            err += 'The individual information of this variant is missing!\n'
            err += 'Please confirm the information in the table submitted and submit again.\n\n'
            continue
        try:
            pmidi = pubinfo.objects.get(pmid=int(pmid))
        except:
            err += 'Error: variant id: '+str(varid)+'\n'
            err += 'The publication information of this variant is missing!\n'
            err += 'Please confirm the information in the table submitted and submit again.\n\n'
        ##Processing Gene information
        try:
            gene_set = geneinfo.objects.filter(entrez=int(entrez))
        except:
            err += 'Error: variant id: '+str(varid)+'\n'
            err += 'The entrez id of this variant is probably depracated or incorrect. Please check it again.\n\n'
        if not gene_set.exists():
            gene_info = get_info(entrez)
            symbol = gene_info.get('symbol','None')
            full_name = gene_info.get('full_name', 'None')
            ids = gene_info.get('ids', 'None')
            location = gene_info.get('location','None')
            names = gene_info.get('names','None')
            summary = gene_info.get('summary', 'None')
            geneitem = geneinfo(entrez=int(entrez),symbol=symbol,fullname=full_name,location=location,other_ids=ids,other_names=names,summary=summary)
            geneitem.save()

        entrezi = geneinfo.objects.get(entrez=int(entrez))

        if not var:
            item = Varinfo(entrez=entrezi, gene=gene, chrom=chrom, start=int(start), end=int(end), dna_ref_nt=dna_ref_nt, dna_alt_nt=dna_alt_nt, hgvs=hgvs, genome_assembly=genome_assembly, exon_intron=exon_intron, exon_number=exon_number, exon_nc=exon_nc, protein_position=intnone(protein_position), pro_ref_aa=pro_ref_aa, pro_alt_aa=pro_alt_aa, frameshift=frameshift, aa_indel=aa_indel, cdna_position=cdna_position, cdna_ref_aa=cdna_ref_aa, cdna_alt_aa=cdna_alt_aa, nt_indel=nt_indel, mrna_accession=mrna_accession, mrna_length=intnone(mrna_length), ref_length=intnone(ref_length), disease=disease)
            item.save()
            item.indid.add(indidi)
            item.pmid.add(pmidi)
            num+=1
        elif not indidi in var.indid.all():
	    var.indid.add(indidi)
            if not pmidi in var.pmid.all():
                var.pmid.add(pmidi)
        var = bool(var) and var or item
        mosaicitem = mosaic(ind=indidi, af_lower=floatnone(af_lower), af_upper=floatnone(af_upper), total_red=intnone(total_red), sample_type=sample_type, method=method,var=var)
        mosaicitem.save()
    return num,err
    
def inpub(f):
    num = 0
    for i in range(len(f)):
        pmid, title, journal, date, disease, population, incidence_lower, incidence_upper, male_cases, female_cases, other_cases, paternal_age_effect = f[i][0:12]
        pub_set = pubinfo.objects.filter(pmid=int(pmid))
        if pub_set.exists():
            continue
        item = pubinfo(pmid=int(pmid), title=title, journal=journal, date=date, disease=disease, population=population, incidence_lower=incidence_lower, incidence_higher=incidence_upper, male_cases=intnone(male_cases), female_cases=intnone(female_cases), other_cases=intnone(other_cases), paternal_age_effect=paternal_age_effect)
        item.save()
        num+=1
    return num

def inind(f):
    sys.stderr.write("Staring insert individual information!\n")
    num = 0
    for i in range(len(f)):
        indid, pmid, whose_mosaic, phenotype_mosaic, disease, omim, patient_mosaic_origin, age_lower, age_upper, affected_child_nc, affected_male_child_nc, affected_female_child_nc, affected_grandson, affected_granddaughter = f[i][0:14]
        sys.stderr.write(str(indid)+'\t'+str(phenotype_mosaic)+str(omim)+'\n')
        ind_set = indinfo.objects.filter(indid=indid)
        if ind_set.exists():
            continue
        sys.stderr.write(str(indid) + str(pmid) + str(whose_mosaic)+str(disease)+str(omim)+'\n')
        affected_female_child_nc=intnone(affected_female_child_nc)
        item = indinfo(indid=indid, pmid=int(pmid), whose_mosaic=whose_mosaic, patient_mosaic_origin=patient_mosaic_origin, phenotype_mosaic=intzero(phenotype_mosaic), age_lower=floatnone(age_lower), age_upper=floatnone(age_upper), affected_child_nc=intnone(affected_child_nc), affected_male_child_nc=intnone(affected_male_child_nc), affected_female_child_nc=intnone(affected_female_child_nc), affected_grandson=intnone(affected_grandson), affected_granddaughter=intnone(affected_granddaughter), disease=disease, omim=intzero(omim))
        item.save()
        num+=1
    return num
	
def insert(f):
    filename = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    filewriter = open('/mnt/e/linux/website/data/update/'+filename+'.xlsx','wb')
    filewriter.write(f.read())
    filewriter.close()
    try:
        xls = load_workbook('/mnt/e/linux/website/data/update/'+filename+'.xlsx')
    except:
        return "Please upload a xlsx format file. Other file formats are not supported for this option!"
    sheetnames = xls.sheetnames
    if not (len(sheetnames)==3 and 'Publication Info' in sheetnames and 'Individual Info' in sheetnames and 'Variation Info' in sheetnames):
        return "The uploaded xlsx is malformed! Please check and re-upload!"
    pubs = read_table(xls,'Publication Info')
    inds = read_table(xls,'Individual Info')
    varis = read_table(xls,'Variation Info')
    numpubs = inpub(pubs[1:])
    numinds = inind(inds[1:])
    numvars,errors = invar(varis[1:])
    sys.stderr.write(errors)
    return "%d publications, %d individuals and %d variants are inserted successfully!\nErrors are shown below:\n%s" %(numpubs, numinds, numvars, errors)
    

def insert_deprecated(f):
    file = ''
    for chunk in f.chunks():
        for line in chunk.split('\n'):
            if len(line.split()) > 0:
                ind, varid, gene, chrom, start, end, met, dis = line.split()
                item = Variant(varid=varid,gene=gene,chrom=chrom,start=start,end=end,disease=dis,method=met)
                item.save()
            else:
                continue
    return "Insert successfully!"

def intnone(str):
    return bool(str) and int(str) or None

def floatnone(str):
    return bool(str) and float(str) or None

def var(f):
    num = 0
    for chunk in f.chunks():
        for line in chunk.split('\n'):
            if len(line.split('\t')) >= 33:
                if line.startswith('V'):
                    continue
                sys.stderr.write(line)
                varid, indid, entrez, pmid, gene, chrom, start, end, dna_ref_nt, dna_alt_nt, hgvs, genome_assembly, exon_intron, exon_number, exon_nc, protein_position, pro_ref_aa, pro_alt_aa, frameshift, aa_indel, cdna_position, cdna_ref_aa, cdna_alt_aa, nt_indel, mrna_accession, mrna_length, ref_length, af_lower, af_upper, total_red, sample_type, method, disease = line.split('\t')[0:33]
                var_set = Varinfo.objects.filter(varid = int(varid))
                indidi = indinfo.objects.get(indid=indid)
                entrezi = geneinfo.objects.get(entrez=int(entrez))
                pmidi = pubinfo.objects.get(pmid=int(pmid))
                if not var_set.exists():
                    item = Varinfo(varid=int(varid), entrez=entrezi, gene=gene, chrom=chrom, start=int(start), end=int(end), dna_ref_nt=dna_ref_nt, dna_alt_nt=dna_alt_nt, hgvs=hgvs, genome_assembly=genome_assembly, exon_intron=exon_intron, exon_number=exon_number, exon_nc=exon_nc, protein_position=intnone(protein_position), pro_ref_aa=pro_ref_aa, pro_alt_aa=pro_alt_aa, frameshift=frameshift, aa_indel=aa_indel, cdna_position=cdna_position, cdna_ref_aa=cdna_ref_aa, cdna_alt_aa=cdna_alt_aa, nt_indel=nt_indel, mrna_accession=mrna_accession, mrna_length=intnone(mrna_length), ref_length=intnone(ref_length), af_lower=floatnone(af_lower), af_upper=floatnone(af_upper), total_red=intnone(total_red), sample_type=sample_type, method=method, disease=disease)
		    item.indid.add(indidi)
		    item.pmid.add(pmidi)
                    item.save()
                    num+=1
                gene_set = geneinfo.objects.filter(entrez=int(entrez))
                if gene_set.exists():
                    continue
                gene_info = get_info(entrez)
                symbol = gene_info.get('symbol','None')
                full_name = gene_info.get('full_name', 'None')
                ids = gene_info.get('ids', 'None')
                location = gene_info.get('location','None')
                names = gene_info.get('names','None')
                summary = gene_info.get('summary', 'None')
                item = geneinfo(entrez=int(entrez),symbol=symbol,fullname=full_name,location=location,other_ids=ids,other_names=names,summary=summary)
                item.save()
            else:
                continue
    return "Insert %d variants successfully!" %(num,)

def pub(f):
    num = 0
    for chunk in f.chunks():
        for line in chunk.split("\n"):
            if len(line.split('\t')) >= 12:
                if line.startswith('P'):
                    continue
                sys.stderr.write(line)
                pmid, title, journal, date, disease, population, incidence_lower, incidence_upper, male_cases, female_cases, other_cases, paternal_age_effect = line.split('\t')[0:12]
                pub_set = pubinfo.objects.filter(pmid=int(pmid))
                if pub_set.exists():
                    continue
                item = pubinfo(pmid=int(pmid), title=title, journal=journal, date=date, disease=disease, population=population, incidence_lower=incidence_lower, incidence_higher=incidence_upper, male_cases=intnone(male_cases), female_cases=intnone(female_cases), other_cases=intnone(other_cases), paternal_age_effect=paternal_age_effect)
                item.save()
                num+=1
            else:
                continue
    return "Insert %d publications successfully!" %(num, )

def ind(f):
    num = 0
    for chunk in f.chunks():
        for line in chunk.split('\n'):
            print len(line.split('\t'))
            if len(line.split('\t')) >= 14:
                if line.startswith('I'):
                    continue
                sys.stderr.write(line)
                indid, pmid, whose_mosaic, patient_mosaic_origin, phenotype_mosaic, age_lower, age_upper, affected_child_nc, affected_male_child_nc, affected_female_child_nc, affected_grandson, affected_granddaughter, disease, omim = line.split('\t')[0:14]
                ind_set = indinfo.objects.filter(indid=indid)
                if ind_set.exists():
                    continue
                item = indinfo(indid=indid, pmid=int(pmid), whose_mosaic=whose_mosaic, patient_mosaic_origin=patient_mosaic_origin, phenotype_mosaic=phenotype_mosaic, age_lower=floatnone(age_lower), age_upper=floatnone(age_upper), affected_child_nc=intnone(affected_child_nc), affected_male_child_nc=intnone(affected_male_child_nc), affected_female_child_nc=intnone(affected_female_child_nc), affected_grandson=intnone(affected_grandson), affected_granddaughter=intnone(affected_granddaughter), disease=disease, omim=int(omim))
                item.save()
                num+=1
            else:
                continue
    return "Insert %d individuals successfully!" %(num, )

